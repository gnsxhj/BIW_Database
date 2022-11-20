#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :ImportProject.py
# @Time      :2020/10/3 14:47
# @Author    :
from PyQt5.QtWidgets import QApplication, QDialog,QMainWindow, QFileDialog
from PyQt5 import uic
import openpyxl
import pandas as pd
import re
class NewProject(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi("gui/LoadNewProject_Dialog.ui")
        self.ui.B_ChooseNewProject.clicked.connect(self.F_ChooseNewProject)
        self.ui.B_Load.clicked.connect(self.F_LoadExcel)
    def F_ChooseNewProject(self):
        s = QFileDialog.getOpenFileName(self, "Open file dialog", "./", "Excel File(*.xlsx)")
        self.Excel = str(s[0])
        self.ui.E_ProjectPath.setText(self.Excel)
        # f = open(path.mainpath+"/newproject.csv", 'w')
        # f.close()
    def F_LoadExcel(self):
        wb = openpyxl.load_workbook(self.Excel)
        sheet_names = wb.sheetnames
        sheet_name = self.ui.E_SheetName.text().strip()
        try:
            for item in sheet_names:
                if sheet_name == item.strip():
                    sheet_name = item #解决Excel sheetname带空格（"abc "） 从而不能识别的问题
                    ws = wb[item]
            pattern_name = re.compile(r'[(VW)(SK)(AU)(AUDI)]',re.I)
            pattern_config = re.compile(r'配置')
            pattern_date = re.compile(r'日期')
            self.ProjectName = re.sub(r'\W','_',self.SearchKeyWord(ws,pattern_name,3).strip())
            self.Configuration = self.SearchKeyWord(ws,pattern_config,3)
            self.Date = self.SearchKeyWord(ws,pattern_date,3)
            self.ui.E_ProjectName.setText(self.ProjectName)
            self.ui.E_Date.setText(self.Date)
            self.ui.E_Configuration.setText(self.Configuration)
            self.colname = list(self.columnname().keys())
            cols = list(self.columnname().values())
            self.newtable_orig = pd.read_excel(self.Excel, sheet_name, header=1,names = self.colname, usecols = cols,
                                          converters = {"ShortPartName":str},skiprows=[2])
            self.newtable = self.newtable_orig.fillna(0)
        except Exception as e:
            print(e)
    def columnname(self):
        attr = dict(
            ShortPartName=self.TextToIndex(self.ui.E_ShortPartName),
            PartName=self.TextToIndex(self.ui.E_PartName),
            DeutschName=self.TextToIndex(self.ui.E_DeutschName),
            CNName=self.TextToIndex(self.ui.E_CNName),
            PlatformKz=self.TextToIndex(self.ui.E_PlatformKz),
            Pfeil=self.TextToIndex(self.ui.E_Pfeil),
            ChildrenAmount=self.TextToIndex(self.ui.E_ChildrenAmount),
            PRNr=self.TextToIndex(self.ui.E_PRNr),
            SetGroup=self.TextToIndex(self.ui.E_Set),
            Class=self.TextToIndex(self.ui.E_Class),
            Amount=self.TextToIndex(self.ui.E_Amount),
            Weight=self.TextToIndex(self.ui.E_Weight),
            Material=self.TextToIndex(self.ui.E_Material),
            Thickness=self.TextToIndex(self.ui.E_Thickness),
            Cladding=self.TextToIndex(self.ui.E_Cladding),
            Strength=self.TextToIndex(self.ui.E_Strength),
            Supplier=self.TextToIndex(self.ui.E_Supplier)
        )
        return attr
    def SearchKeyWord(self,Table,pattern,j):
        for row in Table.iter_rows(min_row=1, max_col=52, max_row=j):
            for cell in row:
                con = str(cell.value).strip()
                item = pattern.search(con)
                if item:
                    return con

    def TextToIndex(self,string):
        index = openpyxl.utils.column_index_from_string(string.text().upper()) - 1
        return index